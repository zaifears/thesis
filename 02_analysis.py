"""
FULL ANALYSIS PIPELINE (v2 — real data, ordinal-first specification)
=======================================================================
Runs everything Chapter 5/6 needs: descriptives, targeted reliability check,
correlation, a PRIMARY ordinal logistic (proportional-odds) regression on the
actual ordered transaction-volume category, a SECONDARY OLS-on-midpoints
model with heteroskedasticity-robust (HC3) SEs for comparison, and the full
diagnostic battery for the OLS model (VIF, normality, heteroskedasticity,
influential points). No statsmodels required -- both models are implemented
by hand with numpy/scipy.

Run this AFTER 02_clean.py has produced coded_dataset.xlsx from the REAL
Google Forms export (not synthetic data).

Every table prints as a clean boxed grid AND is saved to analysis_tables.xlsx,
one table per sheet -- copy a range into Word for a native table.
"""

import pandas as pd
import numpy as np
from scipy import stats
from scipy.optimize import minimize
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_PATH = "coded_dataset.xlsx"
EXCEL_OUT = "analysis_tables.xlsx"
df = pd.read_excel(DATA_PATH)

sheets = {}


def show(title, table_df, floatfmt=".4f"):
    print(f"\n{title}")
    print(tabulate(table_df, headers="keys", tablefmt="fancy_grid",
                    floatfmt=floatfmt, disable_numparse=True,
                    showindex=isinstance(table_df.index, pd.Index)
                    and not isinstance(table_df.index, pd.RangeIndex)))
    sheet_name = title[:31]
    sheets[sheet_name] = table_df.reset_index() if (
        isinstance(table_df.index, pd.Index)
        and not isinstance(table_df.index, pd.RangeIndex)) else table_df.copy()


def fmt_p(p):
    return "< .001" if p < 0.001 else f"{p:.4f}"


print("=" * 78)
print("STEP 0: SAMPLE / DATA QUALITY OVERVIEW")
print("=" * 78)
quality_df = pd.DataFrame([{
    "N": int(len(df)),
    "Flagged (data quality)": int(df["any_quality_flag"].sum()),
    "Flagged %": round(float(df["any_quality_flag"].mean()) * 100, 1),
    "Missing e-KYC (Not sure)": int(df["ekyc_dummy"].isna().sum()),
}]).astype({"N": str, "Flagged (data quality)": str, "Missing e-KYC (Not sure)": str})
show("Table 0. Sample and Data Quality Overview", quality_df, floatfmt=".1f")
print("NOTE: rows are flagged, not silently dropped. Decide and document your")
print("exclusion rule (e.g. drop only if straight-lined AND mismatch flagged)")
print("before finalizing N for the regression models below.")

# ---- STEP 1: Descriptive statistics ----
print("\n" + "=" * 78)
print("STEP 1: DESCRIPTIVE STATISTICS")
print("=" * 78)
for col, label in [("age_bracket", "Age"), ("gender", "Gender"),
                    ("education", "Education"), ("residence", "Residence")]:
    counts = df[col].value_counts()
    pct = (counts / counts.sum() * 100).round(1)
    demo_df = pd.DataFrame({"Category": counts.index, "n": counts.values, "%": pct.values})
    show(f"Table 1{label[0]}. Demographic Profile - {label}", demo_df, floatfmt=".1f")

desc_cols = ["income_bdt", "years_using_mfs", "num_providers", "txn_volume_bdt",
             "txn_frequency", "kyc_friction", "limit_friction",
             "monitoring_perception", "freeze_count", "satisfaction"]
desc_stats = df[desc_cols].describe().T.round(2)
desc_stats.index.name = "Variable"
desc_stats["count"] = desc_stats["count"].astype(int)
show("Table 2. Descriptive Statistics (Continuous/Coded Variables)", desc_stats, floatfmt=".2f")

kyc_counts = df["ekyc_dummy"].value_counts(dropna=False)
print(f"\nKYC method counts (1=e-KYC, 0=Paper, NaN='Not sure'): {kyc_counts.to_dict()}")
if kyc_counts.get(0, 0) < 15 or kyc_counts.get(1, 0) < 15:
    print("WARNING: one KYC-method group has fewer than 15 cases. Interpret the")
    print("e-KYC dummy's coefficient cautiously, or consider dropping it from the model.")

# ---- STEP 2: Reliability (only where a real construct exists) ----
print("\n" + "=" * 78)
print("STEP 2: RELIABILITY CHECK")
print("=" * 78)


def cronbach_alpha(items_df):
    items_df = items_df.dropna()
    k = items_df.shape[1]
    item_vars = items_df.var(axis=0, ddof=1)
    total_var = items_df.sum(axis=1).var(ddof=1)
    return (k / (k - 1)) * (1 - item_vars.sum() / total_var)


friction_items = df[["kyc_friction", "limit_friction"]]
alpha_friction = cronbach_alpha(friction_items)
reliability_df = pd.DataFrame([
    {"Scale": "Compliance friction sub-index (KYC friction + limit friction, "
              "both reverse-coded so higher = more friction)",
     "k": 2, "Cronbach's alpha": round(alpha_friction, 3)},
])
show("Table 3. Reliability Statistics", reliability_df, floatfmt=".3f")
print("NOTE: monitoring_perception and freeze_count are kept as SEPARATE predictors,")
print("not folded into this index -- they measure enforcement intensity, a distinct")
print("concept from KYC/limit friction. Do not force a single combined alpha across")
print("all four; report them as related-but-distinct compliance dimensions instead.")

# ---- STEP 3: Correlation matrix ----
print("\n" + "=" * 78)
print("STEP 3: CORRELATION MATRIX")
print("=" * 78)
corr_cols = ["income_bdt", "years_using_mfs", "num_providers", "kyc_friction",
             "limit_friction", "freeze_count", "monitoring_perception",
             "ekyc_dummy", "txn_volume_bdt", "txn_frequency"]
corr_labels = ["Income", "Yrs MFS", "#Prov", "KYC Friction", "Limit Friction",
               "Freeze", "Monitor", "e-KYC", "Volume", "Freq"]
corr = df[corr_cols].corr(method="pearson").round(2)
corr.columns = corr_labels
corr.index = corr_labels
corr.index.name = "Variable"
show("Table 4. Pearson Correlation Matrix", corr, floatfmt=".2f")

# ============================================================================
# Shared setup: regression-ready frame
# ============================================================================
reg_df = df.dropna(subset=["ekyc_dummy", "txn_volume_cat"]).copy()
AGE_MID = {"18–24": 21, "25–34": 29.5, "35–44": 39.5, "45–54": 49.5,
           "55 or above / ৫৫ বা তার বেশি": 58}
reg_df["age_num"] = reg_df["age_bracket"].map(AGE_MID)
# Alternative if you prefer not to assume within-bracket midpoints:
# age_dummies = pd.get_dummies(reg_df["age_bracket"], prefix="age", drop_first=True)

X_cols = ["kyc_friction", "limit_friction", "freeze_count", "monitoring_perception",
          "ekyc_dummy", "income_bdt", "age_num", "years_using_mfs",
          "num_providers", "urban"]
X_labels = ["KYC Friction", "Limit Friction", "Freeze Count", "Monitoring Perception",
            "e-KYC Dummy", "Income (BDT)", "Age", "Years Using MFS",
            "# Providers", "Urban"]
reg_df = reg_df.dropna(subset=X_cols + ["txn_volume_bdt"])
print(f"\nFinal regression sample after dropping missing predictors: N = {len(reg_df)}")

X_raw = reg_df[X_cols].astype(float).values
X_names = ["Constant"] + X_labels

# ============================================================================
# PRIMARY MODEL: Ordinal logistic (proportional-odds), hand-built
# ============================================================================
print("\n" + "=" * 78)
print("STEP 4a: PRIMARY MODEL -- ORDINAL LOGISTIC REGRESSION")
print("DV = txn_volume_cat (ordered, 1=lowest bracket .. 6=highest bracket)")
print("=" * 78)


class OrdinalLogit:
    """Proportional-odds (cumulative logit) model, no external dependency
    beyond numpy/scipy. Thresholds are reparameterized as
    theta_1 = a[0], theta_j = theta_{j-1} + exp(a[j]) for j>1, which
    guarantees theta_1 < theta_2 < ... without needing constrained optimization."""

    def __init__(self, y, X):
        self.y = np.asarray(y, dtype=int)          # categories coded 1..k
        self.X = np.asarray(X, dtype=float)         # NO intercept column here
        self.n, self.p = self.X.shape
        self.k = int(self.y.max())                  # number of categories
        self.n_thresh = self.k - 1

    def _thresholds(self, a):
        thresh = np.empty(self.n_thresh)
        thresh[0] = a[0]
        for j in range(1, self.n_thresh):
            thresh[j] = thresh[j - 1] + np.exp(a[j])
        return thresh

    def _neg_loglik(self, params):
        a = params[:self.n_thresh]
        beta = params[self.n_thresh:]
        thresh = self._thresholds(a)
        eta = self.X @ beta
        ll = 0.0
        for j in range(1, self.k + 1):
            mask = self.y == j
            if not mask.any():
                continue
            lower = thresh[j - 2] - eta[mask] if j > 1 else -np.inf * np.ones(mask.sum())
            upper = thresh[j - 1] - eta[mask] if j < self.k else np.inf * np.ones(mask.sum())
            p_upper = np.where(np.isinf(upper), 1.0, 1 / (1 + np.exp(-upper)))
            p_lower = np.where(np.isinf(lower), 0.0, 1 / (1 + np.exp(-lower)))
            prob = np.clip(p_upper - p_lower, 1e-10, 1.0)
            ll += np.sum(np.log(prob))
        return -ll

    def fit(self):
        x0 = np.concatenate([
            np.linspace(-1, 1, self.n_thresh),
            np.zeros(self.p),
        ])
        res = minimize(self._neg_loglik, x0, method="BFGS",
                        options={"maxiter": 2000, "gtol": 1e-6})
        self.result = res
        self.params = res.x
        self.thresholds = self._thresholds(self.params[:self.n_thresh])
        self.beta = self.params[self.n_thresh:]
        # numerical Hessian-based SEs
        hess_inv = res.hess_inv
        if hasattr(hess_inv, "todense"):
            hess_inv = np.asarray(hess_inv.todense())
        else:
            hess_inv = np.asarray(hess_inv)
        self.se_all = np.sqrt(np.clip(np.diag(hess_inv), 0, None))
        self.se_beta = self.se_all[self.n_thresh:]
        self.z_beta = self.beta / self.se_beta
        self.p_beta = 2 * (1 - stats.norm.cdf(np.abs(self.z_beta)))
        self.loglik = -res.fun
        return self

    def null_loglik(self):
        null_model = OrdinalLogit(self.y, np.zeros((self.n, 1)))
        null_model.fit()
        return null_model.loglik


y_ord = reg_df["txn_volume_cat"].values
X_scaled = X_raw.copy()
# standardize continuous predictors for optimizer stability; interpret betas
# as effect per 1-SD change (report this explicitly in Chapter 6 tables).
cont_idx = [X_labels.index(l) for l in
            ["Income (BDT)", "Age", "Years Using MFS"]]
scale_factors = {}
for i in cont_idx:
    mu, sd = X_scaled[:, i].mean(), X_scaled[:, i].std()
    scale_factors[X_labels[i]] = (mu, sd)
    X_scaled[:, i] = (X_scaled[:, i] - mu) / sd

ol_model = OrdinalLogit(y_ord, X_scaled).fit()
ll_null = ol_model.null_loglik()
mcfadden_r2 = 1 - (ol_model.loglik / ll_null)
lr_stat = 2 * (ol_model.loglik - ll_null)
lr_p = 1 - stats.chi2.cdf(lr_stat, df=len(X_labels))

ol_summary_df = pd.DataFrame([{
    "N": ol_model.n, "Categories (k)": ol_model.k,
    "Log-likelihood": round(ol_model.loglik, 2),
    "McFadden's R\u00b2": round(mcfadden_r2, 4),
    "LR chi\u00b2": round(lr_stat, 3), "df": len(X_labels), "p": fmt_p(lr_p),
}])
show("Table 5. Ordinal Logistic Model Summary (Primary Model)", ol_summary_df, floatfmt=".4f")

or_values = np.exp(ol_model.beta)
ol_coef_df = pd.DataFrame({
    "Variable": X_labels,
    "B (logit)": np.round(ol_model.beta, 3),
    "SE": np.round(ol_model.se_beta, 3),
    "Odds Ratio": np.round(or_values, 3),
    "z": np.round(ol_model.z_beta, 3),
    "p": [fmt_p(p) for p in ol_model.p_beta],
    "Sig.": ["*" if p < 0.05 else "" for p in ol_model.p_beta],
})
show("Table 6. Ordinal Logistic Coefficients (DV: Txn Volume Category)", ol_coef_df, floatfmt=".3f")
print("Interpretation: OR > 1 for a friction variable means HIGHER friction is")
print("associated with a HIGHER odds of being in a higher-usage category --")
print("i.e. the OPPOSITE of the hypothesized direction. OR < 1 supports H1.")
print("Income/Age/Years-MFS betas are per 1-SD change (standardized for fitting).")

# ============================================================================
# SECONDARY MODEL: OLS on bracket midpoints, with HC3 robust SE
# ============================================================================
print("\n" + "=" * 78)
print("STEP 4b: SECONDARY MODEL -- OLS ON MIDPOINT-CODED VOLUME (ROBUSTNESS CHECK)")
print("DV = txn_volume_bdt (bracket midpoint approximation -- see limitation note)")
print("=" * 78)


class SimpleOLS:
    def __init__(self, y, X):
        self.y = np.asarray(y, dtype=float)
        self.X = np.asarray(X, dtype=float)
        self.n, self.k = self.X.shape
        XtX_inv = np.linalg.inv(self.X.T @ self.X)
        self.XtX_inv = XtX_inv
        self.beta = XtX_inv @ self.X.T @ self.y
        self.fitted = self.X @ self.beta
        self.resid = self.y - self.fitted
        self.dof = self.n - self.k
        self.sse = np.sum(self.resid ** 2)
        self.sigma2 = self.sse / self.dof
        self.cov_beta = self.sigma2 * XtX_inv
        self.se = np.sqrt(np.diag(self.cov_beta))
        self.t = self.beta / self.se
        self.p = 2 * (1 - stats.t.cdf(np.abs(self.t), df=self.dof))
        sst = np.sum((self.y - self.y.mean()) ** 2)
        self.r2 = 1 - self.sse / sst
        self.adj_r2 = 1 - (1 - self.r2) * (self.n - 1) / self.dof
        self.f_stat = (self.r2 / (self.k - 1)) / ((1 - self.r2) / self.dof)
        self.f_p = 1 - stats.f.cdf(self.f_stat, self.k - 1, self.dof)
        self.hat_diag = np.einsum('ij,jk,ik->i', self.X, XtX_inv, self.X)

    def robust_se(self):
        """HC3 heteroskedasticity-consistent standard errors."""
        h = self.hat_diag
        u = self.resid / (1 - h)
        meat = self.X.T @ (self.X * (u ** 2)[:, None])
        cov_hc3 = self.XtX_inv @ meat @ self.XtX_inv
        se_hc3 = np.sqrt(np.diag(cov_hc3))
        t_hc3 = self.beta / se_hc3
        p_hc3 = 2 * (1 - stats.t.cdf(np.abs(t_hc3), df=self.dof))
        return se_hc3, t_hc3, p_hc3


X = np.column_stack([np.ones(len(X_raw)), X_raw])
y = reg_df["txn_volume_bdt"].astype(float).values
model = SimpleOLS(y, X)
se_hc3, t_hc3, p_hc3 = model.robust_se()

model_summary_df = pd.DataFrame([{
    "N": model.n, "R\u00b2": round(model.r2, 4), "Adj. R\u00b2": round(model.adj_r2, 4),
    "F": round(model.f_stat, 3), "df1": model.k - 1, "df2": model.dof, "p": fmt_p(model.f_p),
}])
show("Table 7. OLS Model Summary (Secondary/Robustness)", model_summary_df, floatfmt=".4f")

coef_df = pd.DataFrame({
    "Variable": X_names,
    "B": np.round(model.beta, 3),
    "SE (classical)": np.round(model.se, 3),
    "SE (HC3 robust)": np.round(se_hc3, 3),
    "t (robust)": np.round(t_hc3, 3),
    "p (robust)": [fmt_p(p) for p in p_hc3],
    "Sig.": ["*" if p < 0.05 else "" for p in p_hc3],
})
show("Table 8. OLS Coefficients, Robust SE (DV: Txn Volume BDT, midpoint-coded)", coef_df, floatfmt=".3f")
print("Report the ROBUST (HC3) columns as your primary OLS inference regardless")
print("of what Breusch-Pagan says below -- it costs nothing and protects you either way.")

# ---- STEP 5: Diagnostics (for the OLS model only -- ordinal logit doesn't need these) ----
print("\n" + "=" * 78)
print("STEP 5: DIAGNOSTIC TESTS (OLS Secondary Model)")
print("=" * 78)

vif_rows = []
for i, name in enumerate(X_names):
    if name == "Constant":
        continue
    others = [j for j in range(X.shape[1]) if j != i]
    aux = SimpleOLS(X[:, i], X[:, others])
    vif = 1 / (1 - aux.r2) if aux.r2 < 1 else np.inf
    vif_rows.append({"Variable": name, "VIF": round(vif, 2)})
vif_df = pd.DataFrame(vif_rows)
show("Table 9. Variance Inflation Factors (flag if VIF > 5)", vif_df, floatfmt=".2f")

shapiro_stat, shapiro_p = stats.shapiro(model.resid)
jb_stat, jb_p = stats.jarque_bera(model.resid)
resid_sq = model.resid ** 2
bp_aux = SimpleOLS(resid_sq, X)
bp_lm = model.n * bp_aux.r2
bp_p = 1 - stats.chi2.cdf(bp_lm, df=X.shape[1] - 1)
dw = np.sum(np.diff(model.resid) ** 2) / model.sse
mse = model.sigma2
cooks_d = (model.resid ** 2 / (model.k * mse)) * (model.hat_diag / (1 - model.hat_diag) ** 2)
threshold = 4 / model.n
n_influential = int((cooks_d > threshold).sum())

diagnostics_df = pd.DataFrame([
    {"Test": "Shapiro-Wilk (normality)", "Statistic": round(shapiro_stat, 4), "p": fmt_p(shapiro_p),
     "Verdict": "Normal" if shapiro_p > 0.05 else "Non-normal"},
    {"Test": "Jarque-Bera (normality)", "Statistic": round(jb_stat, 4), "p": fmt_p(jb_p),
     "Verdict": "Normal" if jb_p > 0.05 else "Non-normal"},
    {"Test": "Breusch-Pagan (heteroskedasticity)", "Statistic": round(bp_lm, 4), "p": fmt_p(bp_p),
     "Verdict": "Homoskedastic" if bp_p > 0.05 else "Heteroskedastic -- HC3 SE above already covers this"},
    {"Test": "Durbin-Watson (autocorrelation)", "Statistic": round(dw, 3), "p": "\u2014",
     "Verdict": "Not diagnostic for cross-sectional survey data -- reported for completeness only"},
    {"Test": "Cook's Distance (influential points)", "Statistic": n_influential, "p": "\u2014",
     "Verdict": f"{n_influential} of {model.n} exceed 4/n threshold"},
])
show("Table 10. Diagnostic Test Summary", diagnostics_df, floatfmt=".4f")

# ---- Hypothesis testing summary (uses the PRIMARY ordinal model) ----
print("\n" + "=" * 78)
print("HYPOTHESIS TESTING SUMMARY (based on primary ordinal logistic model)")
print("=" * 78)
hypothesis_map = {
    "KYC Friction": "H1: Higher KYC friction -> lower usage",
    "Limit Friction": "H2: More limit restriction -> lower usage",
    "Freeze Count": "H3: More freezes -> lower usage",
    "Monitoring Perception": "H4: Higher monitoring -> lower usage",
    "e-KYC Dummy": "H5: e-KYC vs. paper -> different usage",
}
hyp_rows = []
for i, name in enumerate(X_labels):
    if name not in hypothesis_map:
        continue
    b, p = ol_model.beta[i], ol_model.p_beta[i]
    direction_ok = b < 0  # negative logit coef = friction lowers odds of higher usage
    hyp_rows.append({
        "Hypothesis": hypothesis_map[name],
        "B (logit)": round(b, 3), "OR": round(np.exp(b), 3), "p": fmt_p(p),
        "Direction": "As hypothesized" if direction_ok else "Opposite of hypothesized",
        "Outcome": "Supported" if (p < 0.05 and direction_ok) else "Not supported",
    })
hypothesis_df = pd.DataFrame(hyp_rows)
show("Table 11. Hypothesis Testing Summary", hypothesis_df, floatfmt=".3f")

# ============================================================================
# Excel export
# ============================================================================
wb = Workbook()
wb.remove(wb.active)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

for sheet_name, table_df in sheets.items():
    ws = wb.create_sheet(title=sheet_name[:31])
    for row in dataframe_to_rows(table_df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(name="Times New Roman", bold=True, size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Times New Roman", size=11)
            cell.alignment = Alignment(horizontal="center")
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) for c in col_cells if c.value is not None)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 3, 10), 45)

wb.save(EXCEL_OUT)

print("\n" + "=" * 78)
print(f"DONE. All tables saved to {EXCEL_OUT}, one per sheet.")
print("Lead with Tables 5-6 (ordinal logit) as your primary Chapter 6 result.")
print("Present Tables 7-8 (OLS + robust SE) as a robustness check, explicitly")
print("noting the midpoint-coding assumption for the open-ended top bracket.")
print("=" * 78)
